'''from imsdatabase import *'''
from stockdatabase import *
from projectdatabase import *
from logdatabase import*
from productdatabase import *
from tkinter import ttk
import tkinter as tk
from customtkinter import *
import customtkinter as ctk
import tkinter.messagebox as tkmb
from tkinter import font
import sqlite3
import openpyxl
from openpyxl import *
from PIL import Image,ImageTk
from CTkTable import CTkTable
from datetime import datetime
from tkinter import filedialog
import pandas as pd
import random
import shutil
wb_bill = openpyxl.load_workbook('Inventory Management System.xlsx')
sheet_bill=wb_bill["Bill Sheet"]
wb = openpyxl.load_workbook('Order Stock.xlsx')
sheet = wb["Order Stock"]
ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("green")
global IMS 
IMS=ctk.CTk()
IMS.geometry("960x540")
IMS.title("Inventory Management System")
IMS.resizable(False, False)
titlefont = ctk.CTkFont(size=44,weight="bold")
normalfont = ctk.CTkFont(size=22)

def login_screen():
    global user_entry
    global user_pass
    global login_screen_frame
    global mainlabel
    mainlabel = ctk.CTkLabel(IMS,text="Inventory Management System",font=titlefont)  
    mainlabel.pack(pady=30)  
    login_screen_frame = ctk.CTkFrame(IMS)  
    login_screen_frame.pack(pady=30,padx=40,fill='both',expand=True)  
    label = ctk.CTkLabel(login_screen_frame,text='Login',font=normalfont,width=200,height=40)  
    label.pack(pady=20,padx=10)  
    user_entry= ctk.CTkEntry(login_screen_frame,placeholder_text="Username",width=200,height=40)  
    user_entry.pack(pady=20,padx=10)  
    user_pass= ctk.CTkEntry(login_screen_frame,placeholder_text="Password",show="*",width=200,height=40)  
    user_pass.pack(pady=20,padx=10)  
    button = ctk.CTkButton(login_screen_frame,text='Login',command=login,width=200,height=40)  
    button.pack(pady=20,padx=10)
    user_pass.bind("<Return>", lambda event: login())

def login():
    global user
    usernameadmin = "A"  
    passwordadmin = "A"
    usernameowner = "Aarav"  
    passwordowner = "IMS"
    usernamestock = "Stock"
    passwordstock = "Stock"
    usernameemployee ="Employee1"
    passwordemployee ="Employee1"
    if user_entry.get() == usernameadmin and user_pass.get() == passwordadmin:
        user="admin"
        mainsrc()
    elif user_entry.get() == usernameowner and user_pass.get() == passwordowner:
        user="owner"
        mainsrc()
    elif user_entry.get() == usernamestock and user_pass.get() == passwordstock:
        user="stock"
        mainsrc()
    elif user_entry.get() == usernameemployee and user_pass.get() == passwordemployee:
        user="employee"
        mainsrc()
    else:  
        tkmb.showerror(title="Login Failed",message="Invalid Username and password")
    #desc=user+" Logged in"
    #log_function_call("Login",user,desc)

def mainsrc():
    mainlabel.destroy()
    login_screen_frame.destroy()
    global navbar
    navbar=ctk.CTkFrame(IMS,width=176, height=650,)
    navbar.pack(fill="y", anchor="w",side="left")
    if user == "admin":
        #button1 = ctk.CTkButton(navbar, text='Bill',command=bills)  
        #button1.grid(row=0,pady=20,padx=70,)
        button2 = ctk.CTkButton(navbar, text='Stock',command=stock)  
        button2.grid(row=1,pady=20,padx=70,)
        button3 = ctk.CTkButton(navbar, text='Project',command=project)
        button3.grid(row=3,pady=20,padx=70,)
        button4 = ctk.CTkButton(navbar, text='Logout',command=logout)
        button4.grid(row=5,pady=20,padx=70,)
        button5= ctk.CTkButton(navbar, text='View Database',command=database_view)
        button5.grid(row=4,pady=20,padx=70,)
        button6= ctk.CTkButton(navbar, text='Product',command=product)
        button6.grid(row=2,pady=20,padx=70,)

    elif user == "owner":
        #button1 = ctk.CTkButton(navbar, text='Bill',command=bills)  
        #button1.grid(row=0,pady=20,padx=70,)
        button2 = ctk.CTkButton(navbar, text='Stock',command=stock)  
        button2.grid(row=1,pady=20,padx=70,)
        button3 = ctk.CTkButton(navbar, text='Project',command=project)
        button3.grid(row=3,pady=20,padx=70,)
        button4 = ctk.CTkButton(navbar, text='Logout',command=logout)
        button4.grid(row=4,pady=20,padx=70,)
        button6= ctk.CTkButton(navbar, text='Product',command=product)
        button6.grid(row=2,pady=20,padx=70,)
    elif user == "stock":
        #button1 = ctk.CTkButton(navbar, text='Bill',command=bills)  
        #button1.grid(row=0,pady=20,padx=70,)
        button2 = ctk.CTkButton(navbar, text='Stock',command=stock)  
        button2.grid(row=1,pady=20,padx=70,)
        button3 = ctk.CTkButton(navbar, text='Project',command=project)
        button3.grid(row=2,pady=20,padx=70,)
        button4 = ctk.CTkButton(navbar, text='Logout',command=logout)
        button4.grid(row=4,pady=20,padx=70,)
        button6= ctk.CTkButton(navbar, text='Product',command=product)
        button6.grid(row=3,pady=20,padx=70,)
    elif user == "employee":
        #button1 = ctk.CTkButton(navbar, text='Bill',command=bills)  
        #button1.grid(row=0,pady=20,padx=70,)
        button2 = ctk.CTkButton(navbar, text='Stock',command=stock)  
        button2.grid(row=1,pady=20,padx=70,)
        button3 = ctk.CTkButton(navbar, text='Project',command=project)
        button3.grid(row=2,pady=20,padx=70,)
        button4 = ctk.CTkButton(navbar, text='Logout',command=logout)
        button4.grid(row=4,pady=20,padx=70,)
        button6= ctk.CTkButton(navbar, text='Product',command=product)
        button6.grid(row=3,pady=20,padx=70,)
'''
def bills():
    clear()
    global bills_frame
    bills_frame = ctk.CTkFrame(IMS)
    bills_frame.pack(pady=20,padx=40,fill='both',expand=True)
    add_bill_button=ctk.CTkButton(bills_frame,text="Add Bill",command=bill_input)
    add_bill_button.grid(row=0,column=0,padx=80,pady=20)
    search_bill_button=ctk.CTkButton(bills_frame,text="Search Bill",command=search_bills)
    search_bill_button.grid(row=0,column=2,padx=80,pady=20)

def bill_input():
    global bill_input_frame
    global bill_input_entry
    global bill_input_entry2
    global bill_input_entry3
    global bill_input_entry4
    global bill_input_entry5
    global bill_input_entry6
    global bill_input_entry7
    global bill_input_entry10
    bill_input_frame = ctk.CTkFrame(bills_frame,width=550,height=410)
    bill_input_frame.grid(row=1,column=0,columnspan=3,pady=(0,20),padx=20)
    bill_input_entry = ctk.CTkEntry(bill_input_frame, placeholder_text="ER:NO",)
    bill_input_entry.grid(row=0,column=0,padx=70,pady=20)
    bill_input_entry2 = ctk.CTkEntry(bill_input_frame, placeholder_text="Item Name",)
    bill_input_entry2.grid(row=1,column=0,padx=70,pady=20)
    bill_input_entry3 = ctk.CTkEntry(bill_input_frame, placeholder_text="Unit",)
    bill_input_entry3.grid(row=2,column=0,padx=70,pady=20)
    bill_input_entry4 = ctk.CTkEntry(bill_input_frame, placeholder_text="Quantity",)
    bill_input_entry4.grid(row=3,column=0,padx=70,pady=20)
    bill_input_entry5 = ctk.CTkEntry(bill_input_frame, placeholder_text="Rate",)
    bill_input_entry5.grid(row=0,column=1,padx=70,pady=20)
    bill_input_entry6 = ctk.CTkEntry(bill_input_frame, placeholder_text="Amount",)
    bill_input_entry6.grid(row=1,column=1,padx=70,pady=20)
    bill_input_entry7 = ctk.CTkEntry(bill_input_frame, placeholder_text="Date DD/MM/YYYY",)
    bill_input_entry7.grid(row=2,column=1,padx=70,pady=20)
    bill_input_entry10 = ctk.CTkEntry(bill_input_frame, placeholder_text="Project Name",)
    bill_input_entry10.grid(row=3,column=1,padx=70,pady=20)
    empty_label = ctk.CTkLabel(bill_input_frame,text="")
    empty_label.grid(row=4,column=0,padx=70,pady=20)
    upload_bill_button=ctk.CTkButton(bill_input_frame,text="Upload Bills",command=upload_bills)
    upload_bill_button.grid(row=4,column=1,padx=30,pady=20)
    button1 = ctk.CTkButton(bill_input_frame, text='Save',command=savebill_input)  
    button1.grid(row=5,column=0,pady=20,padx=70)
    button3 = ctk.CTkButton(bill_input_frame, text='Close',command=closebills)  
    button3.grid(row=5,column=1,pady=20,padx=70)

def upload_bills(): 
    file_path = filedialog.askopenfilename(title="Select file", filetypes=[("Excel files", "*.xlsx")]) 
    if file_path: 
        try:
            df = pd.read_excel(file_path)
            for index, row in df.iterrows(): 
                new_bill( 
                    erno=row['ER.No'], 
                    itemname=row['Item Name'], 
                    unit=row['Unit'], 
                    quantity=row['Quantity'], 
                    rate=row['Rate'], 
                    amount=row['Amount'], 
                    dateday=row['DD'], 
                    datemonth=row['MM'], 
                    dateyear=row['YYYY'], 
                    projectname=row['Project Name'] 
                ) 
            tkmb.showinfo("Success", "Bills uploaded successfully.") 
        except Exception as e:
            tkmb.showerror("Error", f"An error occurred: {e}")
    else: 
        tkmb.showerror("Error", "No file selected.") 
    closebills()
    
def search_bills():
    search_bills_frame1 = ctk.CTkFrame(bills_frame, width=550, height=80)
    search_bills_frame1.grid(row=1, column=0, columnspan=3, pady=0, padx=20)
    search_bills_frame2 = ctk.CTkFrame(bills_frame, width=530, height=370)
    search_bills_frame2.grid(row=2, column=0, columnspan=3, pady=20, padx=20)
    global selected_option
    selected_option = ctk.StringVar(value="Item Name")
    options = ["ER.No", "Item Name", "Unit", "Quantity", "Rate", "Amount", "Date","Month","Year", "Project Name"]
    searchcrit = ctk.CTkOptionMenu(search_bills_frame1, variable=selected_option, values=options)
    searchcrit.grid(row=0, column=0, padx=20, pady=20)
    global search_bill_entry
    search_bill_entry=ctk.CTkEntry(search_bills_frame1, placeholder_text="Search")
    search_bill_entry.grid(row=0, column=1, padx=20, pady=20)
    search_button = ctk.CTkButton(search_bills_frame1, text="Search", command=get_result)
    search_button.grid(row=0, column=2, padx=20, pady=20)
    close_button = ctk.CTkButton(search_bills_frame1, text="Close", command=bills_frame.destroy)
    close_button.grid(row=1, column=2, padx=20, pady=20)
    clear_button=ctk.CTkButton(search_bills_frame1,text="Clear",command=clear_table)
    clear_button.grid(row=1,column=1,padx=20,pady=20)
    Save_button=ctk.CTkButton(search_bills_frame1,text="Save",command=save_search)
    Save_button.grid(row=1,column=0,padx=20,pady=20)
    global tree
    tree = ttk.Treeview(search_bills_frame2, columns=("ER.No", "Item_name", "Unit", "Quantity", "Rate", "Amount", "DD","MM","YYYY", "Project_name"), show='headings')
    tree.heading("ER.No", text="ER.No")
    tree.heading("Item_name", text="Item Name")
    tree.heading("Unit", text="Unit")
    tree.heading("Quantity", text="Quantity")
    tree.heading("Rate", text="Rate")
    tree.heading("Amount", text="Amount")
    tree.heading("DD", text="DD")
    tree.heading("MM", text="MM")
    tree.heading("YYYY", text="YYYY")
    tree.heading("Project_name", text="Project Name")
    tree.column("ER.No", width=60)
    tree.column("Item_name", width=120)
    tree.column("Unit", width=120)
    tree.column("Quantity", width=60)
    tree.column("Rate", width=50)
    tree.column("Amount", width=70)
    tree.column("DD", width=30)
    tree.column("MM", width=30)
    tree.column("YYYY", width=40)
    tree.column("Project_name", width=100)
    tree.pack(fill=ctk.BOTH, expand=True)

def save_search():
    clear_sheet_bills()
    for result in results:
        sheet_bill.append(result)
    wb_bill.save("Inventory Management System.xlsx")

def clear_sheet_bills():
    while sheet_bill.max_row > 1:
        sheet.delete_rows(2)
    wb_bill.save("Inventory Management System.xlsx")

def clear_table():
    for widget in tree.get_children():
        tree.delete(widget)

def get_result():
    for widget in tree.get_children():
        tree.delete(widget)
    selected_criteria = selected_option.get()
    search_term = search_bill_entry.get()
    global results
    results = []
    if selected_criteria == "ER.No":
        results=get_by_erno(search_term)
    elif selected_criteria == "Item Name":
        results=get_by_itemname(search_term)
    elif selected_criteria == "Unit":
        results=get_by_unit(search_term)
    elif selected_criteria == "Date":
        date_input=datetime.strptime(search_term, "%d/%m/%Y")
        day=date_input.day
        month=date_input.month
        year=date_input.year
        results=get_by_date(day,month,year)
    elif selected_criteria == "Month":
        date_input=datetime.strptime(search_term, "%m/%Y")
        month=date_input.month
        year=date_input.year
        results=get_by_month(month,year)
    elif selected_criteria == "Year":
        results=get_by_year(search_term)
    elif selected_criteria == "Project Name":
        results=get_by_projectname(search_term)
    else:
        tkmb.ERROR("Enter valid input")
    for result in results:
        tree.insert("", "end", values=result)

def closebills():
    bills_frame.destroy()

def savebill_input():
    global save_erno
    global save_itemname
    global save_unit
    global save_quantity
    global save_rate
    global save_amount
    global save_day
    global save_month
    global save_year
    global save_projectname
    global bill_input_day
    global bill_input_month
    global bill_input_year
    date_input=datetime.strptime(bill_input_entry7.get(), "%d/%m/%Y")
    bill_input_day=date_input.day
    bill_input_month=date_input.month
    bill_input_year=date_input.year
    save_erno=bill_input_entry.get()
    save_itemname=bill_input_entry2.get()
    save_unit=bill_input_entry3.get()
    save_quantity=bill_input_entry4.get()
    save_rate=bill_input_entry5.get()
    save_amount=bill_input_entry6.get()
    save_day=bill_input_day
    save_month=bill_input_month
    save_year=bill_input_year
    save_projectname=bill_input_entry10.get()
    new_bill(
        save_erno
        ,save_itemname
        ,save_unit
        ,save_quantity
        ,save_rate
        ,save_amount
        ,save_day
        ,save_month
        ,save_year
        ,save_projectname)
    tkmb.showinfo(title="Save Successful",message="You have Saved the bill Successfully")
    closebills()
'''

def stock():
    clear()
    global stock_frame
    stock_frame = ctk.CTkFrame(IMS)
    stock_frame.pack(pady=20,padx=40,fill='both',expand=True)
    button1 = ctk.CTkButton(stock_frame, text='Add',command=add_stock_input)  
    button1.grid(row=0,column=0,pady=20,padx=30)
    button2 = ctk.CTkButton(stock_frame, text='Withdraw',command=withdraw_stock_input)  
    button2.grid(row=0,column=1,pady=20,padx=30)
    button3 = ctk.CTkButton(stock_frame, text='Search',command=search_stock_input)
    button3.grid(row=0,column=2,pady=20,padx=30)
    
def add_stock_input():
    global stock_frame1
    global add_stock_entry
    global add_stock_entry2
    global add_stock_entry3
    global add_stock_entry4
    global add_stock_entry5
    global add_stock_entry6
    stock_frame1 = ctk.CTkFrame(stock_frame,width=550,height=410)
    stock_frame1.grid(columnspan=3)
    code=random_code_generator()
    add_stock_entry = ctk.CTkEntry(stock_frame1, placeholder_text="ID",)
    add_stock_entry.insert(0, str(code))
    add_stock_entry.grid(row=0,column=0,padx=70,pady=20)
    add_stock_entry2 = ctk.CTkEntry(stock_frame1, placeholder_text="Item Name",)
    add_stock_entry2.grid(row=0,column=1,padx=70,pady=20)
    add_stock_entry3 = ctk.CTkEntry(stock_frame1, placeholder_text="Unit",)
    add_stock_entry3.grid(row=1,column=0,padx=70,pady=20)
    add_stock_entry4 = ctk.CTkEntry(stock_frame1, placeholder_text="Quantity",)
    add_stock_entry4.grid(row=1,column=1,padx=70,pady=20)
    add_stock_entry5 = ctk.CTkEntry(stock_frame1, placeholder_text="Minimum Quantity",)
    add_stock_entry5.grid(row=2,column=1,padx=70,pady=20)
    add_stock_entry6 = ctk.CTkEntry(stock_frame1, placeholder_text="Moq",)
    add_stock_entry6.grid(row=2,column=0,padx=70,pady=20)
    empty_label = ctk.CTkLabel(stock_frame1,text="")
    empty_label.grid(row=3,column=0,padx=70,pady=20)
    empty_label1 = ctk.CTkLabel(stock_frame1,text="")
    empty_label1.grid(row=3,column=1,padx=70,pady=20)
    empty_label2 = ctk.CTkLabel(stock_frame1,text="")
    empty_label2.grid(row=4,column=0,padx=70,pady=20)
    upload_stock_button=ctk.CTkButton(stock_frame1,text="Upload Stock",command=upload_stock)
    upload_stock_button.grid(row=4,column=1,padx=30,pady=20)
    Template_stock_button=ctk.CTkButton(stock_frame1,text="Download Template",command=download_template)
    Template_stock_button.grid(row=4,column=0,padx=30,pady=20)
    button1 = ctk.CTkButton(stock_frame1, text='Add',command=add_stock)  
    button1.grid(row=5,column=0,pady=20,padx=70)
    button3 = ctk.CTkButton(stock_frame1, text='Close',command=closestock)  
    button3.grid(row=5,column=1,pady=20,padx=70)    

def download_template():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        shutil.copy('add_stock_template.xlsx', file_path)
        tkmb.showinfo(title="Template Downloaded", message="Edit the template and upload it to add stock.")
    else:
        tkmb.showwarning(title="Download Cancelled", message="Download operation was cancelled")

def random_code_generator():
    random_code=random.randint(100000,999999)
    check_code=get_stock_by_id(random_code)
    if check_code:
        random_code_generator()
    else:
        return random_code

def withdraw_stock_input():
    stock_frame1 = ctk.CTkFrame(stock_frame, width=550, height=410)
    stock_frame1.grid(columnspan=3)
    global withdraw_stock_entry2
    global withdraw_stock_entry4

    withdraw_stock_entry2 = ctk.CTkEntry(stock_frame1, placeholder_text="Item Name",)
    withdraw_stock_entry2.grid(row=0, column=0, padx=70, pady=20)

    withdraw_stock_entry4 = ctk.CTkEntry(stock_frame1, placeholder_text="Quantity",)
    withdraw_stock_entry4.grid(row=0, column=1, padx=70, pady=20)

    empty_label = ctk.CTkLabel(stock_frame1, text="")
    empty_label.grid(row=2, column=0, padx=70, pady=20)
    empty_label1 = ctk.CTkLabel(stock_frame1, text="")
    empty_label1.grid(row=2, column=1, padx=70, pady=20)
    empty_label2 = ctk.CTkLabel(stock_frame1, text="")
    empty_label2.grid(row=3, column=0, padx=70, pady=20)
    empty_label3 = ctk.CTkLabel(stock_frame1, text="")
    empty_label3.grid(row=3, column=1, padx=70, pady=20)
    empty_label4 = ctk.CTkLabel(stock_frame1, text="")
    empty_label4.grid(row=4, column=0, padx=70, pady=20)
    empty_label5 = ctk.CTkLabel(stock_frame1, text="")
    empty_label5.grid(row=4, column=1, padx=70, pady=20)
    empty_label6 = ctk.CTkLabel(stock_frame1, text="")
    empty_label6.grid(row=1, column=0, padx=70, pady=20)
    empty_label7 = ctk.CTkLabel(stock_frame1, text="")
    empty_label7.grid(row=1, column=1, padx=70, pady=20)

    button1 = ctk.CTkButton(stock_frame1, text='Withdraw', command=withdraw_stock)
    button1.grid(row=5, column=0, pady=20, padx=70)
    button3 = ctk.CTkButton(stock_frame1, text='Close', command=closestock)
    button3.grid(row=5, column=1, pady=20, padx=70)

    # Autofill functionality
    large_font = font.Font(family="Segoe UI", size=12)
    listbox_frame = ctk.CTkFrame(stock_frame1, border_color="#333333", bg_color="#333333")
    listbox_frame.grid(row=1, column=0)
    listbox = tk.Listbox(listbox_frame, font=large_font, highlightthickness=0, height=3, width=18, bg='#333333', fg='white', selectbackground='gray', selectforeground='black', border='0')
    listbox.pack(side=tk.LEFT, fill=tk.BOTH)
    scrollbar = tk.Scrollbar(listbox_frame, width=0, highlightthickness=0, bg='#333333', troughcolor='#333333')
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    item_names = get_all_stock_item_names()

    def show_autofill(event):
        listbox.delete(0, tk.END)
        input_text = withdraw_stock_entry2.get()
        if input_text:
            for item in item_names:
                item_name = item
                if input_text.lower() in item_name.lower():
                    listbox.insert(tk.END, item_name)
            if listbox.size() > 0:
                listbox_frame.grid(row=1, column=0)
            else:
                listbox_frame.grid_remove()
        else:
            listbox_frame.grid_remove()

    def select_item(event):
        if listbox.curselection():
            selected_item = listbox.get(listbox.curselection())
            withdraw_stock_entry2.delete(0, tk.END)
            withdraw_stock_entry2.insert(0, selected_item)
            listbox_frame.grid_forget()

    withdraw_stock_entry2.bind("<KeyRelease>", show_autofill)
    withdraw_stock_entry2.bind("<FocusIn>", lambda event: listbox_frame.grid(row=1, column=0))
    listbox.bind("<ButtonRelease-1>", select_item)
    listbox.bind("<Return>", select_item)
    listbox_frame.grid_forget()

def search_stock_input():
    stock_frame1 = ctk.CTkFrame(stock_frame, width=550, height=410)
    stock_frame1.grid(row=1, columnspan=3)
    global search_stock_entry
    global search_stock_entry2
    global button1
    search_stock_entry = ctk.CTkEntry(stock_frame1, placeholder_text="Item Name")
    search_stock_entry.grid(row=0, column=0, padx=70, pady=20)
    search_stock_entry2 = ctk.CTkEntry(stock_frame1, placeholder_text="id")
    search_stock_entry2.grid(row=0, column=1, padx=70, pady=20)
    button1 = ctk.CTkButton(stock_frame1, text='Search', command=display_search_results)
    button1.grid(row=1, column=0, pady=20, padx=70)
    button3 = ctk.CTkButton(stock_frame1, text='Close', command=closestock)
    button3.grid(row=1, column=1, pady=20, padx=70)
    header = [("ID", "Item Name", "Unit", "Quantity")]
    global stock_table_frame
    stock_table_frame = ctk.CTkScrollableFrame(stock_frame, width=540, height=250)
    stock_table_frame.grid(row=2, columnspan=3, pady=20)
    table = CTkTable(master=stock_table_frame, values=header)
    table.pack(expand=True)

    # Autofill functionality
    large_font = font.Font(family="Segoe UI", size=12)
    listbox_frame = ctk.CTkFrame(stock_frame1, border_color="#333333", bg_color="#333333")
    listbox_frame.grid(row=1, column=0)
    listbox = tk.Listbox(listbox_frame, font=large_font, highlightthickness=0, height=3, width=18, bg='#333333', fg='white', selectbackground='gray', selectforeground='black', border='0')
    listbox.pack(side=tk.LEFT, fill=tk.BOTH)
    scrollbar = tk.Scrollbar(listbox_frame, width=0, highlightthickness=0, bg='#333333', troughcolor='#333333')
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    item_names = get_all_stock_item_names()

    def show_autofill(event):
        listbox.delete(0, tk.END)
        input_text = search_stock_entry.get()
        if input_text:
            for item in item_names:
                item_name = item
                if input_text.lower() in item_name.lower():
                    listbox.insert(tk.END, item_name)
            if listbox.size() > 0:
                listbox_frame.grid(row=1, column=0)
                button1.grid_remove()
            else:
                listbox_frame.grid_remove()
                button1.grid()
        else:
            listbox_frame.grid_remove()
            button1.grid()

    def select_item(event):
        if listbox.curselection():
            selected_item = listbox.get(listbox.curselection())
            search_stock_entry.delete(0, tk.END)
            search_stock_entry.insert(0, selected_item)
            listbox_frame.grid_forget()
            button1.grid()

    search_stock_entry.bind("<KeyRelease>", show_autofill)
    search_stock_entry.bind("<FocusIn>", lambda event: (listbox_frame.grid(row=1, column=0), button1.grid_remove()))
    listbox.bind("<ButtonRelease-1>", select_item)
    listbox.bind("<Return>", select_item)
    listbox_frame.grid_forget()

def closestock():
    stock_frame.destroy()

def add_stock():
    global save_stock_id
    global save_stock_itemname
    global save_stock_unit
    global save_stock_quantity
    save_stock_id=add_stock_entry.get()
    save_stock_itemname=add_stock_entry2.get()
    save_stock_unit=add_stock_entry3.get()
    save_stock_quantity=add_stock_entry4.get()
    save_stock_minqty=add_stock_entry5.get()
    save_stock_moq=add_stock_entry6.get()
    result=get_stock_by_itemname(save_stock_itemname)
    result2=get_stock_by_id(save_stock_id)
    if result:
        available_quantity=int (result[0][3])
        new_quantity= available_quantity + int(save_stock_quantity)
        update_stock(save_stock_itemname, new_quantity)
        #desc=user+" added to existing stock id="+result2+" itemname="+result+" quantity="+save_stock_quantity
        #log_function_call("Add Stock",user,desc)
        tkmb.showinfo(title="Save Successful",message="You have Saved the stock Successfully")
    else:
        if result2:
            tkmb.showerror(title="Error", message="This ID already exists. Please enter a unique ID.")
        else:
            new_stock(
                save_stock_id
                ,save_stock_itemname
                ,save_stock_unit
                ,save_stock_quantity
                ,save_stock_minqty
                ,save_stock_moq
                )
            #desc=user+" added new stock id="+save_stock_id+" itemname="+save_stock_itemname+" unit="+save_stock_unit+" quantity="+save_stock_quantity+" minimumquantity="+save_stock_minqty+" moq="+save_stock_moq
            #log_function_call("New Stock",user,desc)
            tkmb.showinfo(title="Save Successful",message="You have Saved the stock Successfully")
    closestock()
    
def withdraw_stock():
    global withdraw_stock_itemname
    global withdraw_stock_quantity
    withdraw_stock_itemname=withdraw_stock_entry2.get()
    withdraw_stock_quantity=withdraw_stock_entry4.get()
    current_stock=get_stock_by_itemname(withdraw_stock_itemname)
    if current_stock:
        if int(withdraw_stock_quantity)<=int(current_stock[0][3]):
            current_quantity = current_stock[0][3]
            new_quantity = current_quantity - int(withdraw_stock_quantity)
            update_stock(withdraw_stock_itemname,new_quantity)
            #desc=user+" withdrew stock id="+current_stock[0][0]+" itemname="+withdraw_stock_itemname+" quantity="+withdraw_stock_quantity
            #log_function_call("Withdraw Stock",user,desc)
            tkmb.showinfo(title="Withdraw Successful",message="You have Withdraw the stock Successfully")
            closestock()
        else:
            tkmb.showerror(title="Error",message="You don't have enough stock to withdraw")
            closestock()
    else:
        tkmb.showerror(title="Error",message="You don't have this item in stock")
        closestock()

def display_search_results():
    for widget in stock_table_frame.winfo_children():
        widget.destroy()

    item_name = search_stock_entry.get()
    id_stock=search_stock_entry2.get()
    if item_name:
        results = get_stock_by_itemname(item_name)
    else:
        results=get_stock_by_id(id_stock)
    
    if results:
        header = ["ID", "Item Name", "Unit", "Quantity"]
        table = CTkTable(master=stock_table_frame, values=[header])
        table.pack(expand=True)

        for result in results:
            table.add_row(result)
    else:
        tkmb.showinfo(title="No Results", message="No stock found.")

def upload_stock(): 
    file_path = filedialog.askopenfilename(title="Select file", filetypes=[("Excel files", "*.xlsx")]) 
    if file_path: 
        try:
            df = pd.read_excel(file_path)
            for index, row in df.iterrows(): 
                print(row['Item Name'])
                if get_stock_by_itemname(row['Item Name']):
                    available_quantity = get_stock_by_itemname(row['Item Name'])[0][3]
                    new_quantity = available_quantity + row['Quantity']
                    update_stock(row['Item Name'], new_quantity)
                else:
                    new_stock( 
                        id=row['ID'], 
                        itemname=row['Item Name'], 
                        unit=row['Unit'],
                        quantity=row['Quantity'],
                        minimumquantity=row['Minimum Quantity'],
                        moq=row['MOQ']
                    )
            tkmb.showinfo("Success", "Stock uploaded successfully.") 
        except Exception as e:
            tkmb.showerror("Error", f"Upload does not match template. Download template and try again. Error: {e}")
    else: 
        tkmb.showerror("Error", "No file selected.") 
    closestock()

def project():
    clear()
    clear_sheet()
    global project_frame
    global item_name_entry
    global item_quantity_entry
    global project_name_entry
    global req_table
    project_frame = ctk.CTkFrame(IMS)
    project_frame.pack(pady=20, padx=20, fill='both', expand=True)
    project_frame1 = ctk.CTkFrame(project_frame)
    project_frame1.grid(pady=20, padx=20, columnspan=2)
    item_name_entry = ctk.CTkEntry(project_frame1, placeholder_text="Item Name")
    item_name_entry.grid(row=0, column=0, padx=30, pady=10)
    item_quantity_entry = ctk.CTkEntry(project_frame1, placeholder_text="Quantity")
    item_quantity_entry.grid(row=0, column=1, padx=30, pady=10)
    project_name_entry = ctk.CTkEntry(project_frame1, placeholder_text="Project Name")
    project_name_entry.grid(row=0, column=2, pady=10, padx=30)
    add_item_button = ctk.CTkButton(project_frame1, text="Add Product", command=add_product_to_project)
    add_item_button.grid(row=1, column=1, pady=10, padx=30)
    load_project_button = ctk.CTkButton(project_frame1, text="Load Project", command=load_project)
    load_project_button.grid(row=1, column=2, pady=10, padx=30)
    close_project_button = ctk.CTkButton(project_frame1, text="Close", command=close_project)
    close_project_button.grid(row=2, column=1, pady=10, padx=30)
    save_order_button = ctk.CTkButton(project_frame1, text="Download Order", command=download_order)
    save_order_button.grid(row=1,column=0, padx=30, pady=10)
    update_project_button = ctk.CTkButton(project_frame1, text="Update Project", command=update_project)
    update_project_button.grid(row=2,column=2, padx=30, pady=10)
    project_req_table_frame = ctk.CTkFrame(project_frame)
    project_req_table_frame.grid(row=2, column=0, pady=(0,20), padx=(20,5), columnspan=3)
    columns = ["Item", "Required Qty", "Available Qty", "Order Qty", "Min Stock", "MOQ", "Order"]
    req_table = ttk.Treeview(project_req_table_frame, columns=columns, show='headings')
    for col in columns:
        req_table.heading(col, text=col)
        req_table.column(col, width=80)
    req_table.pack(fill='both', expand=True)
    
    large_font = font.Font(family="Segoe UI", size=12)
    listbox_frame = ctk.CTkFrame(project_frame1, border_color="#333333", bg_color="#333333")
    listbox_frame.grid(row=1, column=0)
    listbox = tk.Listbox(listbox_frame, font=large_font, highlightthickness=0, height=3, width=18, bg='#333333', fg='white', selectbackground='gray', selectforeground='black', border='0')
    listbox.pack(side=tk.LEFT, fill=tk.BOTH)
    scrollbar = tk.Scrollbar(listbox_frame, width=0, highlightthickness=0, bg='#333333', troughcolor='#333333')
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    item_names = get_all_product_names()
    
    def show_autofill(event):
        listbox.delete(0, tk.END)
        input_text = item_name_entry.get()
        if input_text:
            for item in item_names:
                item_name = item
                if input_text.lower() in item_name.lower():
                    listbox.insert(tk.END, item_name)
            if listbox.size() > 0:
                listbox_frame.grid(row=1, column=0)
                save_order_button.grid_remove()
            else:
                listbox_frame.grid_remove()
                save_order_button.grid()
        else:
            listbox_frame.grid_remove()
            save_order_button.grid()

    def select_item(event):
        if listbox.curselection():
            selected_item = listbox.get(listbox.curselection())
            item_name_entry.delete(0, tk.END)
            item_name_entry.insert(0, selected_item)
            listbox_frame.grid_forget()
            save_order_button.grid()
    
    item_name_entry.bind("<KeyRelease>", show_autofill)
    listbox.bind("<ButtonRelease-1>", select_item)
    listbox.bind("<Return>", select_item)
    listbox_frame.grid_forget()
    
    project_listbox_frame = ctk.CTkFrame(project_frame1, border_color="#333333", bg_color="#333333")
    project_listbox_frame.grid(row=1, column=2)
    project_listbox = tk.Listbox(project_listbox_frame, font=large_font, highlightthickness=0, height=3, width=18, bg='#333333', fg='white', selectbackground='gray', selectforeground='black', border='0')
    project_listbox.pack(side=tk.LEFT, fill=tk.BOTH)
    project_scrollbar = tk.Scrollbar(project_listbox_frame, width=0, highlightthickness=0, bg='#333333', troughcolor='#333333')
    project_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    project_listbox.config(yscrollcommand=project_scrollbar.set)
    project_scrollbar.config(command=project_listbox.yview)
    project_names = get_all_project_names()
    
    def show_project_autofill(event):
        project_listbox.delete(0, tk.END)
        input_text = project_name_entry.get()
        if input_text:
            for project in project_names:
                if input_text.lower() in project.lower():
                    project_listbox.insert(tk.END, project)
            if project_listbox.size() > 0:
                project_listbox_frame.grid(row=1, column=2)
                load_project_button.grid_remove()
            else:
                project_listbox_frame.grid_remove()
                load_project_button.grid()
        else:
            project_listbox_frame.grid_remove()
            load_project_button.grid()
    
    def select_project(event):
        if project_listbox.curselection():
            selected_project = project_listbox.get(project_listbox.curselection())
            project_name_entry.delete(0, tk.END)
            project_name_entry.insert(0, selected_project)
            project_listbox_frame.grid_forget()
            load_project_button.grid()
    
    project_name_entry.bind("<KeyRelease>", show_project_autofill)
    project_listbox.bind("<ButtonRelease-1>", select_project)
    project_listbox.bind("<Return>", select_project)
    project_listbox_frame.grid_forget()

def update_project():
    project_name = project_name_entry.get()
    if project_name:
        clear_project_table(project_name)
        for row in req_table.get_children():
            item_name = req_table.item(row)["values"][0]
            item_quantity = req_table.item(row)["values"][1]
            add_project_item(project_name, item_name, item_quantity)
        tkmb.showinfo(title="Project Updated", message="Project Updated Successfully")
        close_project()
    else:
        tkmb.showerror(title="Input Error", message="Please enter a valid Project name")

def load_project():
    project_name = project_name_entry.get()
    if req_table.get_children():
        choice = tkmb.askokcancel(title="Load Project", message="Do you want to load the project? Unsaved changes will be lost.")
        if choice:
            for row in req_table.get_children():
                req_table.delete(row)
            load_project()                
    else:
        if project_name:
            project_items = get_all_project_items(project_name)
            for item in project_items:
                req_table.insert("", "end", values=(item[0], item[1],"-","-","-","-","-"))
        else:
            tkmb.showerror(title="Input Error", message="Please enter a valid Project name")

def add_product_to_project():
    product_name = item_name_entry.get()
    project_name = project_name_entry.get()
    product_quantity = item_quantity_entry.get()
    if product_name and project_name:
        project_names=get_all_project_names()
        if project_name in project_names:
            choice = tkmb.askokcancel(title="Project Already Exists", message="Project already exists. Do you want to load the project?")
            if choice:
                load_project()
        else:
            try:
                product_quantity = int(item_quantity_entry.get())
            except ValueError:
                tkmb.showerror(title="Input Error", message="Please enter a valid Quantity.")
                return
            products=get_all_product_names()
            if product_name in products:
                create_project_table(project_name)
                product_result = get_all_items(product_name)
                for result in product_result:
                    item_name = result[0]
                    item_quantity = result[1] * product_quantity
                    add_item_row(item_name, item_quantity, project_name)
                tkmb.showinfo(title="Product Added", message="Product Added to Project")
            else:
                tkmb.showerror(title="Product Not Found", message="Product not found")
    else:
        tkmb.showerror(title="Input Error", message="Please enter a valid Product name and a Project")

def check_same(item_name):
    for row in req_table.get_children():
        if req_table.item(row)["values"][0] == item_name:
            return True
    return False

def add_item_row(item_name, item_quantity, project_name):
    if check_same(item_name):
        for row in req_table.get_children():
            if req_table.item(row)["values"][0] == item_name:
                temp_quantity = req_table.item(row)["values"][1]
                final_quantity = temp_quantity + item_quantity
                req_table.delete(row)
                add_item_row(item_name, final_quantity, project_name)
                break
    else:
        result = get_stock_by_itemname(item_name)
        if result:
            available_quantity = result[0][3]
            minimum_quantity = result[0][4]
            moq = result[0][5]
            if available_quantity > item_quantity:
                quantity_left = available_quantity - item_quantity
                if quantity_left >= minimum_quantity:
                    ordering_quantity = 0
                    order = 0
                    update_stock(item_name, quantity_left)
                else:
                    ordering_quantity = minimum_quantity - quantity_left
                    if moq > ordering_quantity:
                        order = moq
                        remaining_quantity = moq - ordering_quantity
                        update_stock(item_name, remaining_quantity + minimum_quantity)
                    else:
                        order = ordering_quantity
                        remaining_quantity = minimum_quantity
                        update_stock(item_name, remaining_quantity)
                req_table.insert("", "end", values=(item_name, item_quantity, available_quantity, '-', minimum_quantity, moq, order))
                if order:
                    add_row_sheet(item_name, result[0][2], order, project_name)
            elif item_quantity > available_quantity:
                ordering_quantity = item_quantity - available_quantity
                if ordering_quantity >= moq:
                    order = ordering_quantity + minimum_quantity
                    update_stock(item_name, minimum_quantity)
                else:
                    remaining_quantity = moq - ordering_quantity
                    if remaining_quantity >= minimum_quantity:
                        order = moq
                        update_stock(item_name, remaining_quantity)
                    else:
                        remaining_order_quantity = minimum_quantity - remaining_quantity
                        order = moq + remaining_order_quantity
                        update_stock(item_name, minimum_quantity)
                req_table.insert("", "end", values=(item_name, item_quantity, available_quantity, ordering_quantity, minimum_quantity, moq, order))
                add_row_sheet(item_name, result[0][2], order, project_name)
            else:
                ordering_quantity = 0
                if moq >= minimum_quantity:
                    order = moq
                    update_stock(item_name, moq)
                else:
                    order = minimum_quantity
                    update_stock(item_name, minimum_quantity)
                req_table.insert("", "end", values=(item_name, item_quantity, available_quantity, ordering_quantity, minimum_quantity, moq, order))
            add_project_item(project_name, item_name,item_quantity)
        else:
            choice = tkmb.askokcancel(title="No Item Found", message=item_name + " not found in the database. Would you like to place the order?.")
            if choice:
                req_table.insert("", "end", values=(item_name, item_quantity, '-', item_quantity, '-', '-', item_quantity))
            add_row_sheet(item_name, None, item_quantity, project_name)

def close_project():
    project_frame.destroy()

def clear_sheet():
    while sheet.max_row > 1:
        sheet.delete_rows(2)
    wb.save("Order Stock.xlsx")

def add_row_sheet(itemname, unit, quantity, projectname):
    data = [itemname, unit, quantity, projectname]
    sheet.append(data)

def download_order():
    wb.save('Order Stock.xlsx')
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        shutil.copy('Order Stock.xlsx', file_path)
        tkmb.showinfo(title="Order Saved", message="Order saved successfully")
        project_frame.destroy()
    else:
        tkmb.showwarning(title="Save Cancelled", message="Save operation was cancelled")

def product():
    clear()
    global product_frame
    global product_tree
    global product_name_entry
    global item_counter
    global add_item_button
    global save_product_button
    item_counter = 0
    product_frame = ctk.CTkFrame(IMS)
    product_frame.pack(pady=20, padx=40, fill='both', expand=True)
    product_name_entry = ctk.CTkEntry(product_frame, placeholder_text="Product Name")
    product_name_entry.grid(row=0, column=1, padx=30, pady=10)
    save_product_button = ctk.CTkButton(product_frame, text="Save", command=save_product)
    save_product_button.grid(row=0, column=2, padx=10, pady=10)
    add_item_button = ctk.CTkButton(product_frame, text="Add item", command=add_item_row_product)
    add_item_button.grid(row=1, column=1, padx=30, pady=10)
    load_product_button = ctk.CTkButton(product_frame, text="Load Product", command=load_product)
    load_product_button.grid(row=1, column=0, padx=30, pady=10)
    close_project_button = ctk.CTkButton(product_frame, text="Close", command=close_product)
    close_project_button.grid(row=1, column=2, pady=10, padx=30)
    columns = ["Item Name", "Quantity"]
    product_tree = ttk.Treeview(product_frame, columns=columns, show='headings')
    product_tree.heading("Item Name", text="Item Name")
    product_tree.heading("Quantity", text="Quantity")
    product_tree.grid(row=2, column=0, columnspan=3, pady=20, padx=20)
    product_tree.bind("<Double-1>", on_double_click)

    # Autofill functionality
    large_font = font.Font(family="Segoe UI", size=12)
    listbox_frame = ctk.CTkFrame(product_frame, border_color="#333333", bg_color="#333333")
    listbox_frame.grid(row=1, column=1)
    listbox = tk.Listbox(listbox_frame, font=large_font, highlightthickness=0, height=3, width=18, bg='#333333', fg='white', selectbackground='gray', selectforeground='black', border='0')
    listbox.pack(side=tk.LEFT, fill=tk.BOTH)
    scrollbar = tk.Scrollbar(listbox_frame, width=0, highlightthickness=0, bg='#333333', troughcolor='#333333')
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    listbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=listbox.yview)
    item_names = get_all_product_names()

    def show_autofill(event):
        listbox.delete(0, tk.END)
        input_text = product_name_entry.get()
        if input_text:
            for item in item_names:
                item_name = item
                if input_text.lower() in item_name.lower():
                    listbox.insert(tk.END, item_name)
            if listbox.size() > 0:
                listbox_frame.grid(row=1, column=1)
                add_item_button.grid_remove()
            else:
                listbox_frame.grid_remove()
                add_item_button.grid()
        else:
            listbox_frame.grid_remove()
            add_item_button.grid()

    def select_item(event):
        if listbox.curselection():
            selected_item = listbox.get(listbox.curselection())
            product_name_entry.delete(0, tk.END)
            product_name_entry.insert(0, selected_item)
            listbox_frame.grid_forget()
            add_item_button.grid()

    product_name_entry.bind("<KeyRelease>", show_autofill)
    listbox.bind("<ButtonRelease-1>", select_item)
    listbox.bind("<Return>", select_item)
    listbox_frame.grid_forget()

def on_double_click(event):
    region = product_tree.identify_region(event.x, event.y)
    if region == "cell":
        current_item = product_tree.selection()[0]
        column = product_tree.identify_column(event.x)
        column_index = int(column.replace("#", "")) - 1 
        current_values = product_tree.item(current_item)["values"]
        item_name = current_values[0]
        quantity = current_values[1]
        edit_window = ctk.CTkToplevel(IMS)
        edit_window.title("Edit Item")
        ctk.CTkLabel(edit_window, text="Item Name:").grid(row=0, column=0, padx=10, pady=10)
        item_name_entry = ctk.CTkEntry(edit_window)
        item_name_entry.grid(row=0, column=1, padx=10, pady=10)
        item_name_entry.insert(0, item_name)
        ctk.CTkLabel(edit_window, text="Quantity:").grid(row=1, column=0, padx=10, pady=10)
        quantity_entry = ctk.CTkEntry(edit_window)
        quantity_entry.grid(row=1, column=1, padx=10, pady=10)
        quantity_entry.insert(0, quantity)
        save_button = ctk.CTkButton(edit_window, text="Save", command=lambda: save_edit(current_item, item_name_entry.get(), quantity_entry.get(), edit_window))
        save_button.grid(row=2, column=1, padx=10, pady=10)
        large_font = font.Font(family="Segoe UI", size=12)
        listbox_frame = ctk.CTkFrame(edit_window, border_color="#333333", bg_color="#333333")
        listbox_frame.grid(row=1, column=1, rowspan=2)
        listbox = tk.Listbox(listbox_frame, font=large_font, width=15, height=5, highlightthickness=0, bg='#242424', fg='white', selectbackground='gray', selectforeground='black', border='0')
        listbox.pack(side=tk.LEFT, fill=tk.BOTH)
        scrollbar = tk.Scrollbar(listbox_frame, width=0, highlightthickness=0, bg='#333333', troughcolor='#333333')
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)
        listbox_frame.grid_remove()
        def show_autofill(event):
            input_text = item_name_entry.get()
            listbox.delete(0, tk.END)
            item_names = getallitem_names()
            for item in item_names:  
                item_name = item[0] if isinstance(item, tuple) else item
                if input_text.lower() in item_name.lower():
                    listbox.insert(tk.END, item_name)
            if listbox.size() > 0:
                listbox_frame.grid()
                quantity_entry.grid_remove()
                save_button.grid_remove()
            else:
                listbox_frame.grid_remove()
                quantity_entry.grid()
                save_button.grid()
        item_name_entry.bind("<KeyRelease>", show_autofill)

        def select_item(event):
            if listbox.curselection():
                selected_item = listbox.get(listbox.curselection())
                item_name_entry.delete(0, tk.END)
                item_name_entry.insert(0, selected_item)
                listbox_frame.grid_remove()
                quantity_entry.grid()
                save_button.grid()
        listbox.bind("<ButtonRelease-1>", select_item)
        delete_button = ctk.CTkButton(edit_window, text="Delete", command=lambda: delete_item_row_product(current_item, edit_window))
        delete_button.grid(row=2, column=0, padx=10, pady=10)

def save_edit(current_item, new_item_name, new_quantity, edit_window):
    product_tree.item(current_item, values=(new_item_name, new_quantity))
    edit_window.destroy()

def delete_item_row_product(current_item, edit_window):
    product_tree.delete(current_item)
    edit_window.destroy()

def add_item_row_product():
    global item_counter
    item_counter = item_counter + 1
    product_tree.insert('', 'end', values=('-', '-'))

def save_product():
    product_name = product_name_entry.get()
    try:
        clear_product_table(product_name)
    except:
        create_product_table(product_name)
    for item in product_tree.get_children():
        item_values = product_tree.item(item)["values"]
        if item_values and len(item_values) == 2:
            try:
                item_name = item_values[0]
                quantity = int(item_values[1])
            except (ValueError, IndexError) as e:
                tkmb.showerror("Error", f"Invalid data in product tree: {e}")
                continue
            add_product_item(product_name, item_name, quantity)
    tkmb.showinfo("Success", "Product saved successfully.")
    close_product()

def load_product():
    product_load_result = get_all_items(product_name_entry.get())
    for result in product_load_result:
        product_tree.insert('', 'end', values=(result[0], result[1]))
        
def close_product():
    product_frame.destroy()

def database_view():
    clear()
    global database_view_frame
    database_view_frame = ctk.CTkFrame(IMS)
    database_view_frame.pack(pady=20,padx=40,fill='both',expand=True)
    dv_stock_button=ctk.CTkButton(database_view_frame,text="Stock",command=dv_stock)
    dv_stock_button.grid(row=0,column=1,padx=80,pady=(30,0))
    empty_label = ctk.CTkLabel(database_view_frame,text="")
    empty_label.grid(row=0,column=0,padx=80,pady=(30,0))
    empty_label = ctk.CTkLabel(database_view_frame,text="")
    empty_label.grid(row=0,column=2,padx=80,pady=(30,0))

'''def dv_bills():
    dv_bills_frame = ctk.CTkFrame(database_view_frame, width=530, height=370)
    dv_bills_frame.grid(row=1, column=0, columnspan=3, pady=20, padx=20)
    bill_tree=ttk.Treeview(dv_bills_frame, columns=("ER.No", "Item_name", "Unit", "Quantity", "Rate", "Amount", "DD","MM","YYYY", "Project_name"), show='headings')
    bill_tree.heading("ER.No", text="ER.No")
    bill_tree.heading("Item_name", text="Item Name")
    bill_tree.heading("Unit", text="Unit")
    bill_tree.heading("Quantity", text="Quantity")
    bill_tree.heading("Rate", text="Rate")
    bill_tree.heading("Amount", text="Amount")
    bill_tree.heading("DD", text="DD")
    bill_tree.heading("MM", text="MM")
    bill_tree.heading("YYYY", text="YYYY")
    bill_tree.heading("Project_name", text="Project Name")
    bill_tree.column("ER.No", width=60)
    bill_tree.column("Item_name", width=120)
    bill_tree.column("Unit", width=120)
    bill_tree.column("Quantity", width=60)
    bill_tree.column("Rate", width=50)
    bill_tree.column("Amount", width=70)
    bill_tree.column("DD", width=30)
    bill_tree.column("MM", width=30)
    bill_tree.column("YYYY", width=40)
    bill_tree.column("Project_name", width=100)
    bill_tree.pack(fill=ctk.BOTH, expand=True)
    for widget in bill_tree.get_children():
        bill_tree.delete(widget)
    result_dv_bills=getallbills()
    for result in result_dv_bills:
        bill_tree.insert('', 'end', values=result)'''

def dv_stock():
    dv_stock_frame = ctk.CTkScrollableFrame(database_view_frame, width=530, height=370)
    dv_stock_frame.grid(row=1, column=0, columnspan=3, pady=20, padx=20)
    header = ["ID", "Item Name", "Unit", "Quantity"]
    stock_table=CTkTable(dv_stock_frame,values=[header])
    stock_table.pack(fill=ctk.BOTH, expand=True)
    stock_result=getallstock()
    for row in stock_result:
        stock_table.add_row(row)

def logout():
    clear()
    navbar.destroy()
    login_screen()

def clear():
    for widget in IMS.winfo_children():
        if widget != navbar:
            widget.destroy()

def get_all_stock_item_names():
    stock_items = getallstock()  
    item_names = [item[1] for item in stock_items]  
    return item_names

login_screen()
IMS.mainloop()